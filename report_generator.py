import io
import datetime
import pandas as pd
import numpy as np
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable, Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from analytics_engine import AnalyticsEngine

class ExecutivePDFReportGenerator:
    @staticmethod
    def _render_trend_chart_image(df: pd.DataFrame) -> tuple:
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt

            all_cols = list(df.columns)
            num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c]) and not pd.api.types.is_bool_dtype(df[c])]

            if not all_cols or not num_cols:
                return None, None

            x_col = all_cols[0]
            y_col = num_cols[0] if num_cols[0] != x_col and len(num_cols) > 1 else (num_cols[0] if len(all_cols) > 1 else all_cols[0])

            trend = AnalyticsEngine.compute_trend_analysis(df, x_col, y_col, aggregation='sum')
            if 'error' in trend or not trend.get('data'):
                return None, None

            x_vals = [str(d['x']) for d in trend['data']]
            y_vals = [d['y'] for d in trend['data']]

            if len(x_vals) > 25:
                step = max(1, len(x_vals) // 20)
                x_vals = x_vals[::step]
                y_vals = y_vals[::step]

            fig, ax = plt.subplots(figsize=(7.0, 2.5), dpi=140)
            ax.plot(x_vals, y_vals, color='#4338ca', marker='o', markersize=4, linewidth=2, label=f'SUM({y_col})')
            ax.fill_between(range(len(x_vals)), y_vals, color='#6366f1', alpha=0.15)

            ax.set_title(f"Visual Data Trend Plot: SUM({y_col}) by {x_col}", fontsize=10, fontweight='bold', color='#1e1b4b', pad=8)
            ax.set_xlabel(x_col, fontsize=8, fontweight='bold', color='#334155')
            ax.set_ylabel(f"SUM of {y_col}", fontsize=8, fontweight='bold', color='#334155')
            
            ax.tick_params(axis='x', rotation=35 if len(x_vals) > 6 else 0, labelsize=7)
            ax.tick_params(axis='y', labelsize=7)
            ax.grid(True, linestyle='--', alpha=0.5, color='#cbd5e1')
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            plt.tight_layout()
            buf = io.BytesIO()
            plt.savefig(buf, format='png', bbox_inches='tight', transparent=False, facecolor='#ffffff')
            plt.close(fig)
            buf.seek(0)

            summary_meta = {
                "x_col": x_col,
                "y_col": y_col,
                "aggregation": trend.get('aggregation', 'sum'),
                "sample_size": trend.get('sample_size', len(df)),
                "total_value": trend.get('total_value', 0),
                "average_value": trend.get('average_value', 0)
            }

            return buf.getvalue(), summary_meta
        except Exception as e:
            print(f"Trend plot render error: {e}")
            return None, None

    @staticmethod
    def _render_demographics_chart_images(df: pd.DataFrame) -> list:
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt

            dem = AnalyticsEngine.compute_demographics(df)
            cat_dists = dem.get('categorical_distributions', {})

            chart_buffers = []
            palette = ['#4338ca', '#10b981', '#f59e0b', '#f43f5e', '#3b82f6', '#8b5cf6', '#ec4899', '#06b6d4']

            for col_name, dist in list(cat_dists.items())[:4]:
                if not dist:
                    continue

                labels = [d['category'][:15] for d in dist[:6]]
                counts = [d['count'] for d in dist[:6]]
                colors_list = palette[:len(labels)]

                fig, ax = plt.subplots(figsize=(3.4, 2.2), dpi=140)
                
                wedges, texts, autotexts = ax.pie(
                    counts,
                    labels=labels,
                    autopct='%1.1f%%',
                    pctdistance=0.75,
                    startangle=140,
                    colors=colors_list,
                    wedgeprops=dict(width=0.45, edgecolor='white', linewidth=1.5),
                    textprops=dict(fontsize=6.5, color='#0f172a')
                )

                for autotext in autotexts:
                    autotext.set_fontsize(6)
                    autotext.set_weight('bold')

                ax.set_title(f"Demographic: {col_name}", fontsize=8.5, fontweight='bold', color='#1e1b4b', pad=6)
                plt.tight_layout()

                buf = io.BytesIO()
                plt.savefig(buf, format='png', bbox_inches='tight', transparent=False, facecolor='#ffffff')
                plt.close(fig)
                buf.seek(0)
                chart_buffers.append((col_name, buf.getvalue()))

            return chart_buffers
        except Exception as e:
            print(f"Demographics chart render error: {e}")
            return []

    @staticmethod
    def generate_pdf_report(df: pd.DataFrame, dataset_name: str, cleaning_history: list) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        
        PRIMARY = colors.HexColor('#1e1b4b')      # Deep Indigo
        ACCENT = colors.HexColor('#4338ca')       # Indigo
        TEXT_COLOR = colors.HexColor('#0f172a')   # Dark Slate
        MUTED_TEXT = colors.HexColor('#475569')   # Slate
        LIGHT_BG = colors.HexColor('#f8fafc')     # Light Slate fill
        BORDER_COLOR = colors.HexColor('#cbd5e1') # Slate border

        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=PRIMARY,
            alignment=0,
            spaceAfter=4
        )

        subtitle_style = ParagraphStyle(
            'ReportSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=13,
            textColor=MUTED_TEXT,
            alignment=0,
            spaceAfter=10
        )

        heading_style = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=12,
            leading=15,
            textColor=ACCENT,
            spaceBefore=12,
            spaceAfter=6
        )

        body_style = ParagraphStyle(
            'ReportBody',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=8.5,
            leading=11,
            textColor=TEXT_COLOR,
            spaceAfter=3
        )

        bullet_style = ParagraphStyle(
            'ReportBullet',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8.5,
            leading=12,
            textColor=TEXT_COLOR,
            leftIndent=10,
            firstLineIndent=-6,
            spaceAfter=3
        )

        table_text = ParagraphStyle(
            'TableText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=TEXT_COLOR
        )

        table_header = ParagraphStyle(
            'TableHeaderText',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.white
        )

        story = []

        # 1. Header Title & Meta Info
        story.append(Paragraph("DataCleanse & Applied Analytics Studio", title_style))
        story.append(Paragraph("Integrated Executive Analytics & Dataset Health Report", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=1.5, color=ACCENT, spaceBefore=0, spaceAfter=10))

        audit = AnalyticsEngine.audit_dataset(df)
        insights = AnalyticsEngine.generate_automated_insights(df)
        demographics = AnalyticsEngine.compute_demographics(df)

        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Metadata Header Table
        meta_data = [
            [Paragraph("<b>Dataset Name:</b>", body_style), Paragraph(str(dataset_name), body_style),
             Paragraph("<b>Report Date:</b>", body_style), Paragraph(now_str, body_style)],
            [Paragraph("<b>Total Rows:</b>", body_style), Paragraph(f"{audit['total_rows']:,}", body_style),
             Paragraph("<b>Total Columns:</b>", body_style), Paragraph(str(audit['total_columns']), body_style)],
            [Paragraph("<b>Data Quality Score:</b>", body_style), Paragraph(f"<b>{audit['quality_score']}%</b>", body_style),
             Paragraph("<b>Missing Cells:</b>", body_style), Paragraph(f"{audit['total_missing_cells']:,}", body_style)]
        ]
        meta_table = Table(meta_data, colWidths=[110, 160, 110, 160])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), LIGHT_BG),
            ('BOX', (0,0), (-1,-1), 1, BORDER_COLOR),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 10))

        # 2. Key Insights Section
        story.append(Paragraph("Automated Data Insights & Observations", heading_style))
        if insights:
            for ins in insights:
                clean_ins = ins.replace('**', '').replace('[Demographics]', '').replace('[Health]', '').replace('[Data Loss Risk]', '').strip()
                story.append(Paragraph(f"• {clean_ins}", bullet_style))
        else:
            story.append(Paragraph("No specific data anomalies detected.", body_style))
        story.append(Spacer(1, 10))

        # 3. Visual Data Trend Plot Analytics & Trend Summary Card
        trend_bytes, trend_summary = ExecutivePDFReportGenerator._render_trend_chart_image(df)
        if trend_bytes:
            story.append(Paragraph("Visual Data Trend Plot Analytics & Trend Summary", heading_style))

            if trend_summary:
                trend_meta_data = [
                    [
                        Paragraph(f"<b>X-Axis Attribute:</b> {trend_summary['x_col']}", body_style),
                        Paragraph(f"<b>Y-Axis Metric:</b> {trend_summary['y_col']}", body_style),
                        Paragraph(f"<b>Strategy:</b> {trend_summary['aggregation'].upper()}", body_style)
                    ],
                    [
                        Paragraph(f"<b>Records Analyzed:</b> {trend_summary['sample_size']:,}", body_style),
                        Paragraph(f"<b>Aggregated Total:</b> {trend_summary['total_value']:,}", body_style),
                        Paragraph(f"<b>Average Metric:</b> {trend_summary['average_value']:,}", body_style)
                    ]
                ]
                trend_meta_table = Table(trend_meta_data, colWidths=[175, 175, 170])
                trend_meta_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#eef2ff')),
                    ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#c7d2fe')),
                    ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e0e7ff')),
                    ('PADDING', (0,0), (-1,-1), 5),
                ]))
                story.append(trend_meta_table)
                story.append(Spacer(1, 6))

            img_trend = Image(io.BytesIO(trend_bytes), width=515, height=185)
            story.append(img_trend)
            story.append(Spacer(1, 10))

        # 4. Demographic Distributions & Categorical Breakdown Charts
        demo_charts = ExecutivePDFReportGenerator._render_demographics_chart_images(df)
        if demo_charts:
            story.append(Paragraph("Demographic Distributions & Categorical Breakdown", heading_style))
            grid_cells = []
            row_cells = []
            for col_name, chart_img_bytes in demo_charts:
                img_demo = Image(io.BytesIO(chart_img_bytes), width=250, height=160)
                row_cells.append(img_demo)
                if len(row_cells) == 2:
                    grid_cells.append(row_cells)
                    row_cells = []
            if row_cells:
                if len(row_cells) == 1:
                    row_cells.append(Paragraph("", body_style))
                grid_cells.append(row_cells)

            demo_table = Table(grid_cells, colWidths=[255, 255])
            demo_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('PADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(demo_table)
            story.append(Spacer(1, 10))

        # 5. Column Data Health Audit Table
        story.append(Paragraph("Column Data Health Audit", heading_style))
        col_headers = [Paragraph("Column", table_header), Paragraph("Type", table_header),
                       Paragraph("Total Rows", table_header), Paragraph("Missing", table_header), Paragraph("Missing %", table_header)]
        col_rows = [col_headers]
        for col_info in audit.get('columns', []):
            col_rows.append([
                Paragraph(str(col_info['column']), table_text),
                Paragraph(str(col_info['dtype']), table_text),
                Paragraph(f"{audit['total_rows']:,}", table_text),
                Paragraph(f"{col_info['missing_count']:,}", table_text),
                Paragraph(f"{col_info['missing_pct']}%", table_text)
            ])
        col_table = Table(col_rows, colWidths=[150, 90, 100, 100, 100])
        col_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PRIMARY),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(col_table)
        story.append(Spacer(1, 10))

        # 6. Numerical Demographic Metrics Summary
        num_summary = demographics.get('numeric_summary', [])
        if num_summary:
            story.append(Paragraph("Numerical Demographic Metrics Summary", heading_style))
            num_headers = [
                Paragraph("Attribute", table_header), Paragraph("Mean", table_header),
                Paragraph("Std Dev", table_header), Paragraph("Median", table_header),
                Paragraph("Min", table_header), Paragraph("Max", table_header), Paragraph("Skewness", table_header)
            ]
            num_rows = [num_headers]
            for row in num_summary:
                num_rows.append([
                    Paragraph(str(row['column']), table_text),
                    Paragraph(str(row['mean']), table_text),
                    Paragraph(str(row['std']), table_text),
                    Paragraph(str(row['median']), table_text),
                    Paragraph(str(row['min']), table_text),
                    Paragraph(str(row['max']), table_text),
                    Paragraph(str(row['skewness']), table_text)
                ])
            num_table = Table(num_rows, colWidths=[120, 70, 70, 70, 70, 70, 70])
            num_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), ACCENT),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
                ('PADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(num_table)
            story.append(Spacer(1, 10))

        # 7. Data Cleaning History Audit Trail
        if cleaning_history:
            story.append(Paragraph("Data Cleaning Audit Trail History", heading_style))
            for item in cleaning_history:
                story.append(Paragraph(f"• {item}", bullet_style))
            story.append(Spacer(1, 10))

        # 8. Sample Records Table Preview (First 20 Rows)
        story.append(Paragraph("Dataset Sample Records Preview (First 20 Rows)", heading_style))
        sample_df = df.head(20)
        display_cols = list(sample_df.columns[:6])
        
        sample_headers = [Paragraph(str(c)[:15], table_header) for c in display_cols]
        sample_rows = [sample_headers]

        col_w = max(40, min(140, int(540 / max(1, len(display_cols)))))
        col_widths = [col_w] * len(display_cols)

        for _, row in sample_df.iterrows():
            row_cells = [Paragraph(str(row[c])[:25] if pd.notna(row[c]) else "<i>null</i>", table_text) for c in display_cols]
            sample_rows.append(row_cells)

        sample_table = Table(sample_rows, colWidths=col_widths)
        sample_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
            ('PADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(sample_table)

        doc.build(story)
        return buffer.getvalue()


class InventoryReportGenerator:
    """Generates professional executive PDF and Excel reports for the Inventory Management System."""

    @staticmethod
    def generate_inventory_pdf(items: list, store_name: str = "All Items", active_stores_count: int = 0) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=30,
            rightMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        styles = getSampleStyleSheet()

        PRIMARY = colors.HexColor('#0f172a')       # Slate 900
        ACCENT_TEAL = colors.HexColor('#0d9488')   # Teal 600
        INDIGO = colors.HexColor('#4f46e5')        # Indigo 600
        AMBER = colors.HexColor('#d97706')         # Amber 600
        EMERALD = colors.HexColor('#059669')       # Emerald 600
        ROSE = colors.HexColor('#e11d48')          # Rose 600
        TEXT_COLOR = colors.HexColor('#0f172a')
        MUTED_TEXT = colors.HexColor('#64748b')
        LIGHT_BG = colors.HexColor('#f8fafc')
        BORDER_COLOR = colors.HexColor('#e2e8f0')

        title_style = ParagraphStyle(
            'InvReportTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=16,
            leading=20,
            textColor=PRIMARY,
            alignment=0,
            spaceAfter=2
        )

        subtitle_style = ParagraphStyle(
            'InvReportSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8.5,
            leading=12,
            textColor=MUTED_TEXT,
            alignment=0,
            spaceAfter=8
        )

        section_heading = ParagraphStyle(
            'InvSectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=11,
            leading=14,
            textColor=INDIGO,
            spaceBefore=8,
            spaceAfter=5
        )

        table_header = ParagraphStyle(
            'InvTableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=7.5,
            leading=9.5,
            textColor=colors.white,
            alignment=1
        )

        table_text = ParagraphStyle(
            'InvTableText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=7.5,
            leading=9.5,
            textColor=TEXT_COLOR
        )

        table_text_bold = ParagraphStyle(
            'InvTableTextBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=7.5,
            leading=9.5,
            textColor=TEXT_COLOR
        )

        table_text_center = ParagraphStyle(
            'InvTableTextCenter',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=7.5,
            leading=9.5,
            textColor=TEXT_COLOR,
            alignment=1
        )

        table_text_hand = ParagraphStyle(
            'InvTableTextHand',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=EMERALD,
            alignment=1
        )

        table_text_store = ParagraphStyle(
            'InvTableTextStore',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=AMBER,
            alignment=1
        )

        table_text_total = ParagraphStyle(
            'InvTableTextTotal',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=INDIGO,
            alignment=1
        )

        badge_instock = ParagraphStyle(
            'InvBadgeInStock',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=7,
            leading=9,
            textColor=EMERALD,
            alignment=1
        )

        badge_lowstock = ParagraphStyle(
            'InvBadgeLowStock',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=7,
            leading=9,
            textColor=ROSE,
            alignment=1
        )

        kpi_title_style = ParagraphStyle(
            'InvKpiTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=7,
            leading=8.5,
            textColor=MUTED_TEXT,
            alignment=1
        )

        kpi_val_style = ParagraphStyle(
            'InvKpiVal',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=13,
            leading=16,
            textColor=PRIMARY,
            alignment=1
        )

        story = []

        # 1. Header Banner
        story.append(Paragraph("DataCleanse &bull; Inventory Stock Report", title_style))
        gen_time = datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")
        story.append(Paragraph(f"Scope: <b>{store_name}</b> &nbsp;|&nbsp; Generated on: {gen_time}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=1.5, color=ACCENT_TEAL, spaceBefore=2, spaceAfter=8))

        # 2. KPI Summary Cards
        total_items_count = len(items)
        sum_on_hand = sum(int(it.get('qty_on_hand') or 0) for it in items)
        sum_on_store = sum(int(it.get('qty_on_store') or 0) for it in items)
        sum_total_pcs = sum_on_hand + sum_on_store

        kpi_data = [
            [
                Paragraph("ACTIVE STORES", kpi_title_style),
                Paragraph("CATALOG ITEMS", kpi_title_style),
                Paragraph("TOTAL ON HAND (MAIN)", kpi_title_style),
                Paragraph("TOTAL ON STORE (ALLOCATED)", kpi_title_style),
                Paragraph("TOTAL PCS (STOCK)", kpi_title_style),
            ],
            [
                Paragraph(str(active_stores_count if active_stores_count > 0 else (1 if store_name != "All Items" else 0)), kpi_val_style),
                Paragraph(str(total_items_count), kpi_val_style),
                Paragraph(str(sum_on_hand), kpi_val_style),
                Paragraph(str(sum_on_store), kpi_val_style),
                Paragraph(str(sum_total_pcs), kpi_val_style),
            ]
        ]
        kpi_table = Table(kpi_data, colWidths=[110, 110, 110, 110, 112])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), LIGHT_BG),
            ('BOX', (0,0), (-1,-1), 1, BORDER_COLOR),
            ('INNERGRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 10))

        # 3. Inventory Table
        story.append(Paragraph(f"Inventory Stock Breakdown ({store_name})", section_heading))

        headers = [
            Paragraph("Item Name", table_header),
            Paragraph("On Hand<br/><font size='6' color='#cbd5e1'>Main Inventory</font>", table_header),
            Paragraph("On Store<br/><font size='6' color='#cbd5e1'>Allocated</font>", table_header),
            Paragraph("Total Pcs<br/><font size='6' color='#cbd5e1'>Hand + Store</font>", table_header),
            Paragraph("Stock Alert", table_header),
            Paragraph("Last Updated", table_header),
        ]
        table_rows = [headers]

        LOW_STOCK_THRESHOLD = 10

        for it in items:
            name = str(it.get('name') or '—')
            on_hand = int(it.get('qty_on_hand') or 0)
            on_store = int(it.get('qty_on_store') or 0)
            total_pcs = int(it.get('total_pcs') or (on_hand + on_store))
            last_up = str(it.get('last_updated') or '—')[:16]

            # Store breakdown text if available
            store_bd = it.get('store_breakdown')
            if isinstance(store_bd, list) and store_bd:
                bd_text = ", ".join([f"{sb.get('name', 'Store')}: {sb.get('qty', 0)}" for sb in store_bd])
                store_display = Paragraph(f"<b>{on_store}</b><br/><font size='6' color='#64748b'>{bd_text}</font>", table_text_center)
            else:
                store_display = Paragraph(str(on_store), table_text_store)

            is_low = total_pcs <= LOW_STOCK_THRESHOLD
            alert_cell = Paragraph("&#9888; Low Stock" if is_low else "&#10003; In Stock", badge_lowstock if is_low else badge_instock)

            table_rows.append([
                Paragraph(name, table_text_bold),
                Paragraph(str(on_hand), table_text_hand),
                store_display,
                Paragraph(str(total_pcs), table_text_total),
                alert_cell,
                Paragraph(last_up, table_text_center),
            ])

        # Summary Totals Row
        table_rows.append([
            Paragraph("<b>TOTALS</b>", table_text_bold),
            Paragraph(f"<b>{sum_on_hand}</b>", table_text_hand),
            Paragraph(f"<b>{sum_on_store}</b>", table_text_store),
            Paragraph(f"<b>{sum_total_pcs}</b>", table_text_total),
            Paragraph("", table_text),
            Paragraph("", table_text),
        ])

        inv_table = Table(table_rows, colWidths=[200, 70, 100, 65, 60, 57])
        inv_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PRIMARY),
            ('GRID', (0,0), (-1,-2), 0.5, BORDER_COLOR),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, LIGHT_BG]),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#e2e8f0')),
            ('LINEABOVE', (0,-1), (-1,-1), 1.5, PRIMARY),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 3.5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(inv_table)

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def generate_inventory_excel(items: list, store_name: str = "All Items") -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="64748B")
        font_kpi_label = Font(name="Segoe UI", size=8, bold=True, color="475569")
        font_kpi_val = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
        font_header = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
        font_data = Font(name="Segoe UI", size=9, color="0F172A")
        font_data_bold = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
        font_hand = Font(name="Segoe UI", size=9, bold=True, color="047857")
        font_store = Font(name="Segoe UI", size=9, bold=True, color="B45309")
        font_total = Font(name="Segoe UI", size=9, bold=True, color="4338CA")
        font_in_stock = Font(name="Segoe UI", size=8.5, bold=True, color="059669")
        font_low_stock = Font(name="Segoe UI", size=8.5, bold=True, color="DC2626")

        fill_banner = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_kpi = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_totals = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        double_bottom_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='0F172A'),
            bottom=Side(style='double', color='0F172A')
        )

        # 1. Title Banner
        ws.merge_cells("A1:G1")
        c1 = ws["A1"]
        c1.value = "  DataCleanse — Inventory Management & Stock Report"
        c1.font = font_title
        c1.fill = fill_banner
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 32

        # 2. Subtitle info
        gen_time = datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")
        ws["A2"] = f"Scope: {store_name}  |  Exported on: {gen_time}  |  Definition: Total Pcs = Item on Hand (Main) + Item on Store (Allocated)"
        ws["A2"].font = font_subtitle
        ws.row_dimensions[2].height = 18

        # 3. KPI Summary Blocks (Row 4 & 5)
        sum_on_hand = sum(int(it.get('qty_on_hand') or 0) for it in items)
        sum_on_store = sum(int(it.get('qty_on_store') or 0) for it in items)
        sum_total_pcs = sum_on_hand + sum_on_store

        kpis = [
            ("A", "B", "TOTAL CATALOG ITEMS", len(items)),
            ("C", "C", "TOTAL ON HAND (MAIN)", sum_on_hand),
            ("D", "D", "TOTAL ON STORE (ALLOCATED)", sum_on_store),
            ("E", "G", "TOTAL PCS (OVERALL STOCK)", sum_total_pcs),
        ]
        for col_start, col_end, label, val in kpis:
            cell_range_lbl = f"{col_start}4:{col_end}4"
            cell_range_val = f"{col_start}5:{col_end}5"
            if col_start != col_end:
                ws.merge_cells(cell_range_lbl)
                ws.merge_cells(cell_range_val)
            ws[f"{col_start}4"] = label
            ws[f"{col_start}4"].font = font_kpi_label
            ws[f"{col_start}4"].fill = fill_kpi
            ws[f"{col_start}4"].alignment = align_center
            ws[f"{col_start}5"] = val
            ws[f"{col_start}5"].font = font_kpi_val
            ws[f"{col_start}5"].fill = fill_kpi
            ws[f"{col_start}5"].alignment = align_center

            for row in range(4, 6):
                for col_idx in range(openpyxl.utils.column_index_from_string(col_start), openpyxl.utils.column_index_from_string(col_end) + 1):
                    ws.cell(row=row, column=col_idx).border = thin_border

        ws.row_dimensions[4].height = 14
        ws.row_dimensions[5].height = 20

        # 4. Table Headers (Row 7)
        headers = [
            ("A", "SKU"),
            ("B", "Item Name"),
            ("C", "On Hand (Main)"),
            ("D", "On Store (Allocated)"),
            ("E", "Total Pcs"),
            ("F", "Store Breakdown / Allocations"),
            ("G", "Stock Alert"),
        ]
        for col, h in headers:
            cell = ws[f"{col}7"]
            cell.value = h
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[7].height = 24

        # 5. Data Rows
        row_idx = 8
        LOW_STOCK_THRESHOLD = 10

        for it in items:
            sku = str(it.get('sku') or '—')
            name = str(it.get('name') or '—')
            on_hand = int(it.get('qty_on_hand') or 0)
            on_store = int(it.get('qty_on_store') or 0)
            total_pcs = int(it.get('total_pcs') or (on_hand + on_store))

            store_bd = it.get('store_breakdown')
            if isinstance(store_bd, list) and store_bd:
                bd_text = "; ".join([f"{sb.get('name', 'Store')}: {sb.get('qty', 0)}" for sb in store_bd])
            else:
                bd_text = "—"

            is_low = total_pcs <= LOW_STOCK_THRESHOLD
            alert_text = "Low Stock" if is_low else "In Stock"

            ws[f"A{row_idx}"] = sku
            ws[f"A{row_idx}"].font = font_data
            ws[f"A{row_idx}"].alignment = align_center

            ws[f"B{row_idx}"] = name
            ws[f"B{row_idx}"].font = font_data_bold
            ws[f"B{row_idx}"].alignment = align_left

            ws[f"C{row_idx}"] = on_hand
            ws[f"C{row_idx}"].font = font_hand
            ws[f"C{row_idx}"].alignment = align_right

            ws[f"D{row_idx}"] = on_store
            ws[f"D{row_idx}"].font = font_store
            ws[f"D{row_idx}"].alignment = align_right

            ws[f"E{row_idx}"] = total_pcs
            ws[f"E{row_idx}"].font = font_total
            ws[f"E{row_idx}"].alignment = align_right

            ws[f"F{row_idx}"] = bd_text
            ws[f"F{row_idx}"].font = font_data
            ws[f"F{row_idx}"].alignment = align_wrap

            ws[f"G{row_idx}"] = alert_text
            ws[f"G{row_idx}"].font = font_low_stock if is_low else font_in_stock
            ws[f"G{row_idx}"].alignment = align_center

            for col_letter in ["A","B","C","D","E","F","G"]:
                c = ws[f"{col_letter}{row_idx}"]
                c.border = thin_border
                if row_idx % 2 == 0:
                    c.fill = fill_zebra

            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

        # 6. Totals Row
        tot_row = row_idx
        ws[f"A{tot_row}"] = "TOTALS"
        ws[f"A{tot_row}"].font = font_data_bold
        ws[f"A{tot_row}"].alignment = align_center

        ws[f"B{tot_row}"] = f"{len(items)} Items"
        ws[f"B{tot_row}"].font = font_data_bold
        ws[f"B{tot_row}"].alignment = align_left

        ws[f"C{tot_row}"] = sum_on_hand
        ws[f"C{tot_row}"].font = font_hand
        ws[f"C{tot_row}"].alignment = align_right

        ws[f"D{tot_row}"] = sum_on_store
        ws[f"D{tot_row}"].font = font_store
        ws[f"D{tot_row}"].alignment = align_right

        ws[f"E{tot_row}"] = sum_total_pcs
        ws[f"E{tot_row}"].font = font_total
        ws[f"E{tot_row}"].alignment = align_right

        ws[f"F{tot_row}"] = ""
        ws[f"G{tot_row}"] = ""

        for col_letter in ["A","B","C","D","E","F","G"]:
            c = ws[f"{col_letter}{tot_row}"]
            c.border = double_bottom_border
            c.fill = fill_totals

        ws.row_dimensions[tot_row].height = 22

        # 7. Column Auto-fit
        col_widths = {
            "A": 16,
            "B": 36,
            "C": 18,
            "D": 22,
            "E": 16,
            "F": 42,
            "G": 16,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

